import re
import json
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import pandas as pd
import os

from openpyxl.styles.borders import Border, Side

# def read_file(file_path):
#     try:
#         # Parse the XML file
#         tree = ET.parse(file_path)
#         root = tree.getroot()

#         # Find the document_content element
#         document_content = root.find('.//document_content')

#         if document_content is not None:
#             return document_content.text
#         else:
#             print("Error: Could not find document_content in the XML file.")
#             return None
#     except ET.ParseError as e:
#         print(f"Error parsing XML file: {e}")
#         return None
#     except FileNotFoundError:
#         print(f"Error: File not found: {file_path}")
#         return None
# def extract_and_concatenate(input_string):
#     # Extract the part that starts with 'M' followed by digits
#     m_part_match = re.search(r'M\d+', input_string)
#
#     # Extract the numbers after the underscore
#     underscore_part_match = re.search(r'_(\d+)', input_string)
#
#     if m_part_match and underscore_part_match:
#         m_part = m_part_match.group()
#         underscore_part = underscore_part_match.group(1)
#         return m_part + underscore_part
#     else:
#         return "Invalid input format"

def extract_and_concatenate(input_string):
    # Extract the part that starts with 'M' followed by digits
    m_part_match = re.search(r'M\d+', input_string)

    # Extract the number between any two alphabetic characters (with or without an underscore after the first character)
    num_between_alpha_match = re.search(r'[A-Za-z]_?(\d+)(?=[A-Za-z])', input_string)

    if m_part_match and num_between_alpha_match:
        m_part = m_part_match.group()
        num_between_alpha = num_between_alpha_match.group(1)
        return m_part+num_between_alpha
    else:
        return "Invalid input format", "Invalid input format"


def extract_data(content):
    # Extract L and CM value
    l_pattern = r'L=(\d+M\s+\d+\.\d+CM)'
    l_match = re.search(l_pattern, content)
    l_value = l_match.group(1) if l_match else "Not found"

    # Extract U value
    u_pattern = r'U=(\d+\.\d+%)'
    u_match = re.search(u_pattern, content)
    u_value = u_match.group(1) if u_match else "Not found"

    # Extract PERIM value
    perim_pattern = r'PERIM=(\d+\.\d+CM)'
    perim_match = re.search(perim_pattern, content)
    perim_value = perim_match.group(1) if perim_match else "Not found"
    perim_value_wihtout_CM = perim_value.replace('CM','')

    # Extract LBMK value
    lbmk_pattern = r'LBMK:([\w-]+)'
    lbmk_match = re.search(lbmk_pattern, content)
    lbmk_value = lbmk_match.group(1) if lbmk_match else "Not found"

    length_M = l_value.split(' ')[0].replace('M', '')
    length_CM = l_value.split(' ')[1].replace('CM', '')
    length = (int(length_M)) + (float(length_CM)/100)
    lbmk = extract_and_concatenate(lbmk_value)

    return {
        "L": length,
        "U": u_value,
        "PERIM": perim_value_wihtout_CM,
        "LBMK": lbmk,
        "LBMK_full" : lbmk_value
    }

# Main function to tie it all together


def append_to_excel(data):
    # Convert the data dictionary to a pandas DataFrame
    new_df = pd.DataFrame([data])
    excel_path = r'C:\Users\admin\Desktop\accurmarks\Test2.xlsx'
    if os.path.exists(excel_path):
        # If file exists, read it and append new data
        existing_df = pd.read_excel(excel_path)
        updated_df = pd.concat([existing_df, new_df], ignore_index=True)
    else:
        # If file doesn't exist, create new DataFrame
        updated_df = new_df

    # Write the updated DataFrame to Excel
    updated_df.to_excel(excel_path, index=False)
    print(f"Data appended to {excel_path}")


def change_extension(file_path, new_extension):
    base_name, _ = os.path.splitext(file_path)
    new_file_path = base_name + "." + new_extension
    os.rename(file_path, new_file_path)
    return new_file_path
# def find_and_edit_excel(data,appended_items,ws):
#     # Define the column to search and the substring
#     header_row = 1
# # Populate headers-to-columns.
#     fields = {}
#     for cnum in range(1, ws.max_column + 1):
#         field = ws.cell(row=header_row, column=cnum).value
#         fields[field] = cnum
#     substring = data['LBMK']  # Substring to search for
#     percentage_style = NamedStyle(name='percentage_style', number_format='0.00%')
#     number_style = NamedStyle(name='number_style', number_format='0.00')
#     pattern = re.compile(r'^LN')
#     linning = bool(pattern.match(data['LBMK_full']))
#     # Iterate through the cells in the column
#     for row_num in range(header_row + 1, ws.max_row + 1):
#         job = ws.cell(row=row_num, column=fields['Job #']).value
#         cut = ws.cell(row=row_num, column=fields['Cut #']).value
#         concatenate = ws.cell(row=row_num, column=fields['Concatenate']).value
#         job_cut = str(job) + str(cut)
#
#         unique = job_cut + str(data['L']) +str(data['U'])
#
#         # if (linning ==True) and ("Lining" in concatenate):
#         #     pass
#         # else:
#         #     continue
#         if (substring == job_cut) and (row_num not in appended_items):  # Check if the substring is in the cell's value
#             # Edit cells in the same row
#             if (linning ==True):
#                 if ("Lining" in concatenate):
#                     pass
#                 else:
#                     continue
#             ws.cell(row=row_num, column=fields['Marker Length']).value = data['L']  # Modify cell B in the found row
#             ws.cell(row=row_num, column=fields['Marker Length']).style= number_style
#             ws.cell(row=row_num, column=fields['Marker Utilization']).value= data['U']
#             ws.cell(row=row_num, column=fields['Marker Utilization']).style  = percentage_style
#             ws.cell(row=row_num, column=fields['PARAMETER']).value= data['PERIM']
#             ws.cell(row=row_num, column=fields['PARAMETER']).style  = number_style # Modify cell C in the found row
#             appended_items.append(row_num)
#             break  # Exit loop once the substring is found (remove if you want to find multiple occurrences)
#     return row_num
def evaluate_formula(ws, formula):
    # Remove '=CONCATENATE(' and ')'
    formula = formula[len('=CONCATENATE('):-1]
    references = formula.split(',')

    # Extract values from the referenced cells
    values = []
    for ref in references:
        ref = ref.strip()  # Clean up any whitespace
        cell_value = ws[ref].value
        values.append(cell_value)

    # Concatenate the values
    return ''.join(str(value) for value in values)


def find_and_edit_excel(data, appended_items, ws):
    # Define the column to search and the substring
    header_row = 1
    # Populate headers-to-columns.
    fields = {}
    for cnum in range(1, ws.max_column + 1):
        field = ws.cell(row=header_row, column=cnum).value
        fields[field] = cnum

    substring = data['LBMK']  # Substring to search for

    percentage_style = NamedStyle(name='percentage_style', number_format='0.00%')
    number_style = NamedStyle(name='number_style', number_format='0.00')

    # Define border style
    border_style = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Determine if LBMK_full starts with 'LN'
    linning = bool(data['LBMK_full'].startswith('LN'))

    # Iterate through the rows
    for row_num in range(header_row + 1, ws.max_row + 1):
        job = ws.cell(row=row_num, column=fields['Job #']).value
        cut = ws.cell(row=row_num, column=fields['Cut #']).value
        concatenate = ws.cell(row=row_num, column=fields['Concatenate']).value
        job_cut = str(job) + str(cut)

        unique = job_cut + str(data['L']) + str(data['U'])

        # Evaluate formula if necessary
        if concatenate and isinstance(concatenate, str) and concatenate.startswith('=CONCATENATE'):
            concatenated_value = evaluate_formula(ws, concatenate)
        else:
            concatenated_value = concatenate

        if substring == job_cut and row_num not in appended_items:
            # If LBMK_full starts with 'LN', only update if 'Lining' is in concatenated_value
            if linning:
                if "Lining" in str(concatenated_value):
                    # Edit cells in the same row
                    ws.cell(row=row_num, column=fields['Marker Length']).value = data['L']
                    ws.cell(row=row_num, column=fields['Marker Length']).style = number_style
                    ws.cell(row=row_num, column=fields['Marker Utilization']).value = data['U']
                    ws.cell(row=row_num, column=fields['Marker Utilization']).style = percentage_style
                    ws.cell(row=row_num, column=fields['PARAMETER']).value = data['PERIM']
                    ws.cell(row=row_num, column=fields['PARAMETER']).style = number_style
                    # Apply border to the updated cells
                    ws.cell(row=row_num, column=fields['Marker Length']).border = border_style
                    ws.cell(row=row_num, column=fields['Marker Utilization']).border = border_style
                    ws.cell(row=row_num, column=fields['PARAMETER']).border = border_style
                    appended_items.append(row_num)
                    print(f"Data updated in row {row_num}.")
                    break  # Exit loop once the substring is found
            else:
                # If LBMK_full does not start with 'LN', update regardless of 'Lining'
                ws.cell(row=row_num, column=fields['Marker Length']).value = data['L']
                ws.cell(row=row_num, column=fields['Marker Length']).style = number_style
                ws.cell(row=row_num, column=fields['Marker Utilization']).value = data['U']
                ws.cell(row=row_num, column=fields['Marker Utilization']).style = percentage_style
                ws.cell(row=row_num, column=fields['PARAMETER']).value = data['PERIM']
                ws.cell(row=row_num, column=fields['PARAMETER']).style = number_style
                # Apply border to the updated cells
                ws.cell(row=row_num, column=fields['Marker Length']).border = border_style
                ws.cell(row=row_num, column=fields['Marker Utilization']).border = border_style
                ws.cell(row=row_num, column=fields['PARAMETER']).border = border_style
                appended_items.append(row_num)
                print(f"Data updated in row {row_num}.")
                break  # Exit loop once the substring is found

    return row_num
# def main(directory_path, excel_file_path):
#     # Load the workbook
#     try:
#         wb = load_workbook(excel_file_path)
#         print(f"Workbook '{excel_file_path}' loaded successfully.")
#     except Exception as e:
#         print(f"Error loading workbook: {e}")
#         return
#
#     data_json = {}
#     appended_items = []
#     # Select the active sheet
#     try:
#         sheet = wb['Sheet1']
#         print("Sheet 'Sheet1' selected successfully.")
#     except KeyError:
#         print("Error: Sheet 'Sheet1' not found.")
#         return
#     except Exception as e:
#         print(f"Error selecting sheet: {e}")
#         return
#
#     for index, filename in enumerate(os.listdir(directory_path)):
#         print(f'=>Processing {index}. {filename}')
#         file_path = os.path.join(directory_path, filename)
#         if os.path.isfile(file_path):
#             file_path = change_extension(file_path, 'txt')
#             try:
#                 with open(file_path, 'r') as file:
#                     content = file.read()
#             except Exception as e:
#                 print(f"Error reading file {file_path}: {e}")
#                 continue
#
#             if content:
#                 extracted_data = extract_data(content)
#                 data_json[f'{filename}'] = extracted_data
#                 for key, value in extracted_data.items():
#                     print(f"{key}: {value}")
#                 insert = find_and_edit_excel(extracted_data, appended_items, sheet)
#                 extracted_data['unique_id'] = insert
#             else:
#                 print("Failed to extract data due to file reading error.")
#
#     try:
#         with open('log.json', 'w') as json_logs:
#             json_logs.write(json.dumps(data_json, indent=4))
#         print("Log file 'log.json' written successfully.")
#     except Exception as e:
#         print(f"Error writing log file: {e}")
#
#     # Save the updated workbook with a new name
#     try:
#         new_excel_path = f'C:\\Users\\admin\\Desktop\\accurmarks\\updated_{os.path.basename(excel_file_path)}'
#         wb.save(new_excel_path)
#         print(f"Workbook saved successfully as '{new_excel_path}'.")
#     except Exception as e:
#         print(f"Error saving workbook: {e}")

def main(directory_path, excel_file_path):
    # Load the workbook
    try:
        wb = load_workbook(excel_file_path)
        print(f"Workbook '{excel_file_path}' loaded successfully.")
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return None

    data_json = {}
    appended_items = []

    # Select the active sheet
    try:
        sheet = wb['Sheet1']
        print("Sheet 'Sheet1' selected successfully.")
    except KeyError:
        print("Error: Sheet 'Sheet1' not found.")
        return None
    except Exception as e:
        print(f"Error selecting sheet: {e}")
        return None

    for index, filename in enumerate(os.listdir(directory_path)):
        print(f'=>Processing {index}. {filename}')
        file_path = os.path.join(directory_path, filename)
        if os.path.isfile(file_path):
            # Ensure file has a .txt extension before processing
            if not file_path.lower().endswith('.txt'):
                file_path = change_extension(file_path, 'txt')
            try:
                with open(file_path, 'r') as file:
                    content = file.read()
            except Exception as e:
                print(f"Error reading file {file_path}: {e}")
                continue

            if content:
                extracted_data = extract_data(content)
                data_json[f'{filename}'] = extracted_data
                for key, value in extracted_data.items():
                    print(f"{key}: {value}")
                insert = find_and_edit_excel(extracted_data, appended_items, sheet)
                extracted_data['unique_id'] = insert
            else:
                print("Failed to extract data due to file reading error.")

    # Save the updated workbook
    try:
        current_directory = os.getcwd()
        new_excel_path = os.path.join(current_directory, f'updated_{os.path.basename(excel_file_path)}')
        wb.save(new_excel_path)
        print(f"Workbook saved successfully as '{new_excel_path}'.")
    except Exception as e:
        print(f"Error saving workbook: {e}")
        return None

    # Write the log file
    try:
        with open('log.json', 'w') as json_logs:
            json_logs.write(json.dumps(data_json, indent=4))
        print("Log file 'log.json' written successfully.")
    except Exception as e:
        print(f"Error writing log file: {e}")

    print("Processing completed successfully.")
    return new_excel_path

# if __name__ == "__main__":
#     directory_path = r'C:\Users\admin\Desktop\accurmarks\files\files'
#     excel_file_path = r'C:\Users\admin\Desktop\accurmarks\Test2.xlsx'  # Update this path
#     main(directory_path, excel_file_path)